import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import PurePath
import re
from typing import Any
from zipfile import BadZipFile, ZipFile

import numpy as np
import pandas as pd


UNSUPPORTED_FILE_ERROR = "Unsupported file type. Please upload a CSV, TSV, TXT, or XLSX file."
PARSE_FILE_ERROR = (
    "Could not read the uploaded file. Please check that it is a valid "
    "CSV, TSV, TXT, or Excel file."
)
INVALID_TABLE_ERROR = "Uploaded file does not contain a valid table."
TEXT_EXTENSIONS = {".csv", ".tsv", ".txt"}
SUPPORTED_EXTENSIONS = TEXT_EXTENSIONS | {".xlsx"}
DELIMITER_CANDIDATES = ",\t;|"
SNIFF_SAMPLE_SIZE = 64 * 1024
MAX_UPLOAD_BYTES = 50 * 1024 * 1024
MAX_ROWS = 1_000_000
MAX_COLUMNS = 1_000
MAX_COLUMN_NAME_LENGTH = 512
MAX_XLSX_UNCOMPRESSED_BYTES = 200 * 1024 * 1024
MAX_XLSX_ARCHIVE_ENTRIES = 10_000
SAFE_INTEGER_LIMIT = 2**53 - 1
MANUAL_ENCODINGS = {
    "utf-8",
    "utf-8-sig",
    "utf-16",
    "utf-16-le",
    "utf-16-be",
    "utf-32",
    "utf-32-le",
    "utf-32-be",
    "cp1252",
    "latin-1",
    "gb18030",
    "shift_jis",
}
ID_NAME_TOKENS = {
    "id",
    "identifier",
    "mrn",
    "code",
    "barcode",
    "postcode",
    "postal",
    "zip",
}
NUMERIC_TOKEN = re.compile(
    r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?$"
)


@dataclass(frozen=True)
class DatasetMetadata:
    file_name: str
    file_size_bytes: int
    format_name: str
    delimiter: str | None
    encoding: str | None
    sheet_name: str | None = None
    available_sheets: tuple[str, ...] = ()


@dataclass(frozen=True)
class DatasetLoadResult:
    dataframe: pd.DataFrame
    metadata: DatasetMetadata


def read_dataset(
    uploaded_file: Any,
    *,
    encoding_override: str | None = None,
    sheet_name: str | None = None,
) -> pd.DataFrame:
    """Read an uploaded supported file into a DataFrame."""
    return read_dataset_with_metadata(
        uploaded_file,
        encoding_override=encoding_override,
        sheet_name=sheet_name,
    ).dataframe


def read_dataset_with_metadata(
    uploaded_file: Any,
    *,
    encoding_override: str | None = None,
    sheet_name: str | None = None,
) -> DatasetLoadResult:
    """Read a dataset and report the detected file characteristics."""
    file_name = str(getattr(uploaded_file, "name", ""))
    extension = PurePath(file_name).suffix.lower()

    if extension not in SUPPORTED_EXTENSIONS:
        raise ValueError(UNSUPPORTED_FILE_ERROR)

    content = _uploaded_file_bytes(uploaded_file)
    if not content:
        raise ValueError(INVALID_TABLE_ERROR)
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValueError(
            f"Uploaded file is too large. Maximum supported size is "
            f"{MAX_UPLOAD_BYTES // (1024 * 1024)} MB."
        )

    try:
        if extension == ".xlsx":
            _validate_xlsx_archive(content)
            with pd.ExcelFile(BytesIO(content), engine="openpyxl") as workbook:
                available_sheets = tuple(workbook.sheet_names)
                selected_sheet = sheet_name or (
                    available_sheets[0] if available_sheets else None
                )
                if selected_sheet not in available_sheets:
                    raise ValueError(
                        f"Worksheet '{selected_sheet}' does not exist in this workbook."
                    )
                dataframe = pd.read_excel(
                    workbook,
                    sheet_name=selected_sheet,
                    dtype=object,
                )
            dataframe = _restore_excel_formulas(
                dataframe,
                content,
                selected_sheet,
            )
            dataframe = _coerce_object_dataframe(dataframe)
            delimiter = None
            encoding = None
        else:
            if encoding_override is not None and encoding_override not in MANUAL_ENCODINGS:
                raise ValueError(f"Unsupported text encoding '{encoding_override}'.")
            encoding = encoding_override or _detect_text_encoding(content)
            text = content.decode(encoding, errors="strict")
            _validate_text_content(text)
            delimiter = _detect_delimiter(text, extension)
            parse_delimiter = delimiter if delimiter is not None else ","
            _validate_delimited_shape(text, parse_delimiter)
            dataframe = pd.read_csv(
                StringIO(text),
                sep=parse_delimiter,
                dtype=str,
                keep_default_na=False,
                na_filter=False,
                index_col=False,
            )
            dataframe = _coerce_object_dataframe(dataframe)
            available_sheets = ()
            selected_sheet = None
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(PARSE_FILE_ERROR) from exc

    _validate_dataframe(dataframe)

    return DatasetLoadResult(
        dataframe=dataframe,
        metadata=DatasetMetadata(
            file_name=file_name,
            file_size_bytes=len(content),
            format_name=_format_name(extension),
            delimiter=delimiter,
            encoding=encoding,
            sheet_name=selected_sheet,
            available_sheets=available_sheets,
        ),
    )


def _uploaded_file_bytes(uploaded_file: Any) -> bytes:
    if hasattr(uploaded_file, "getvalue"):
        content = uploaded_file.getvalue()
    else:
        if hasattr(uploaded_file, "seek"):
            uploaded_file.seek(0)
        content = uploaded_file.read()

    if isinstance(content, str):
        return content.encode("utf-8")
    if isinstance(content, memoryview):
        return content.tobytes()
    if isinstance(content, bytes):
        return content
    return bytes(content)


def _detect_text_encoding(content: bytes) -> str:
    if content.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    if content.startswith((b"\xff\xfe\x00\x00", b"\x00\x00\xfe\xff")):
        return "utf-32"
    if content.startswith((b"\xff\xfe", b"\xfe\xff")):
        return "utf-16"

    if _looks_like_utf16(content, little_endian=True):
        return "utf-16-le"
    if _looks_like_utf16(content, little_endian=False):
        return "utf-16-be"

    try:
        content.decode("utf-8")
    except UnicodeDecodeError:
        try:
            content.decode("cp1252")
        except UnicodeDecodeError:
            return "latin-1"
        return "cp1252"
    return "utf-8"


def _looks_like_utf16(content: bytes, *, little_endian: bool) -> bool:
    sample = content[:4096]
    if len(sample) < 4:
        return False

    null_positions = sample[1::2] if little_endian else sample[0::2]
    other_positions = sample[0::2] if little_endian else sample[1::2]
    null_ratio = null_positions.count(0) / len(null_positions)
    other_null_ratio = other_positions.count(0) / len(other_positions)
    return null_ratio >= 0.3 and other_null_ratio < 0.1


def _detect_delimiter(text: str, extension: str) -> str | None:
    if extension == ".tsv":
        return "\t"

    sample = text[:SNIFF_SAMPLE_SIZE]
    try:
        return csv.Sniffer().sniff(
            sample,
            delimiters=DELIMITER_CANDIDATES,
        ).delimiter
    except csv.Error:
        return _default_delimiter(extension)


def _default_delimiter(extension: str) -> str | None:
    if extension == ".csv":
        return ","
    if extension == ".tsv":
        return "\t"
    return None


def _validate_text_content(text: str) -> None:
    if "\x00" in text:
        raise ValueError("Uploaded text file contains binary null bytes.")
    sample = text[:SNIFF_SAMPLE_SIZE]
    if sample:
        disallowed = sum(
            ord(character) < 32 and character not in "\n\r\t"
            for character in sample
        )
        if disallowed / len(sample) > 0.01:
            raise ValueError("Uploaded file appears to contain binary data.")


def _validate_xlsx_archive(content: bytes) -> None:
    try:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_XLSX_ARCHIVE_ENTRIES:
                raise ValueError("Excel workbook contains too many archive entries.")
            uncompressed_size = sum(entry.file_size for entry in entries)
            if uncompressed_size > MAX_XLSX_UNCOMPRESSED_BYTES:
                raise ValueError(
                    "Excel workbook expands beyond the supported 200 MB limit."
                )
            for entry in entries:
                if (
                    entry.file_size > 10 * 1024 * 1024
                    and entry.compress_size > 0
                    and entry.file_size / entry.compress_size > 1_000
                ):
                    raise ValueError(
                        "Excel workbook has a suspicious compression ratio."
                    )
    except BadZipFile as exc:
        raise ValueError("Uploaded XLSX file is not a valid Excel archive.") from exc


def _validate_delimited_shape(text: str, delimiter: str) -> None:
    reader = csv.reader(StringIO(text), delimiter=delimiter)
    meaningful_rows: list[tuple[int, list[str]]] = []
    for line_number, row in enumerate(reader, start=1):
        if not row or all(cell == "" for cell in row):
            continue
        meaningful_rows.append((line_number, row))
        if len(meaningful_rows) > MAX_ROWS + 1:
            raise ValueError(
                f"Uploaded table exceeds the maximum of {MAX_ROWS:,} rows."
            )

    if not meaningful_rows:
        raise ValueError(INVALID_TABLE_ERROR)

    _, header = meaningful_rows[0]
    if len(header) > MAX_COLUMNS:
        raise ValueError(
            f"Uploaded table exceeds the maximum of {MAX_COLUMNS:,} columns."
        )
    normalized_headers = [str(value).strip() for value in header]
    if any(not value for value in normalized_headers):
        raise ValueError("Every column must have a non-empty header.")
    if any(len(value) > MAX_COLUMN_NAME_LENGTH for value in normalized_headers):
        raise ValueError(
            f"Column names cannot exceed {MAX_COLUMN_NAME_LENGTH} characters."
        )
    duplicate_headers = sorted(
        {
            value
            for value in normalized_headers
            if normalized_headers.count(value) > 1
        }
    )
    if duplicate_headers:
        raise ValueError(
            "Duplicate column headers are not supported: "
            + ", ".join(duplicate_headers[:10])
        )

    expected_width = len(header)
    for line_number, row in meaningful_rows[1:]:
        if len(row) != expected_width:
            raise ValueError(
                f"Row {line_number} has {len(row)} fields; expected "
                f"{expected_width}. Fix the delimiter or malformed row."
            )


def _validate_dataframe(dataframe: pd.DataFrame) -> None:
    if dataframe.empty or len(dataframe.columns) == 0:
        raise ValueError(INVALID_TABLE_ERROR)
    if len(dataframe) > MAX_ROWS:
        raise ValueError(f"Uploaded table exceeds the maximum of {MAX_ROWS:,} rows.")
    if len(dataframe.columns) > MAX_COLUMNS:
        raise ValueError(
            f"Uploaded table exceeds the maximum of {MAX_COLUMNS:,} columns."
        )

    names = [str(column).strip() for column in dataframe.columns]
    if any(not name or name.lower().startswith("unnamed:") for name in names):
        raise ValueError("Every column must have a non-empty header.")
    if len(set(names)) != len(names):
        raise ValueError("Duplicate column headers are not supported.")
    if any(len(name) > MAX_COLUMN_NAME_LENGTH for name in names):
        raise ValueError(
            f"Column names cannot exceed {MAX_COLUMN_NAME_LENGTH} characters."
        )
    dataframe.columns = names
    non_blank = dataframe.apply(
        lambda series: series.notna()
        & series.astype("string").str.strip().ne("").fillna(False)
    )
    if not bool(non_blank.to_numpy().any()):
        raise ValueError(INVALID_TABLE_ERROR)


def _coerce_object_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    """Infer safe numeric columns without damaging identifier-like text."""
    result = dataframe.copy()
    for column in result.columns:
        series = result[column]
        if not pd.api.types.is_object_dtype(series.dtype):
            continue
        if _is_identifier_column(str(column)):
            result[column] = series.map(
                lambda value: "" if value is None else str(value)
            )
            continue

        strings = series.map(lambda value: "" if value is None else str(value).strip())
        non_empty = strings[strings.ne("")]
        if non_empty.empty or not non_empty.map(
            lambda value: NUMERIC_TOKEN.fullmatch(value) is not None
        ).all():
            continue
        if non_empty.map(_has_lossy_numeric_shape).any():
            continue

        numeric = pd.to_numeric(strings.replace("", pd.NA), errors="coerce")
        finite = numeric.dropna().map(np.isfinite).all()
        if not finite:
            continue
        if numeric.dropna().map(lambda value: float(value).is_integer()).all():
            if numeric.isna().any():
                result[column] = numeric.astype("Int64")
            else:
                result[column] = numeric.astype("int64")
        else:
            result[column] = numeric.astype("float64")
    return result


def _restore_excel_formulas(
    dataframe: pd.DataFrame,
    content: bytes,
    sheet_name: str,
) -> pd.DataFrame:
    """Preserve formula text instead of silently importing uncached formulas as NA."""
    from openpyxl import load_workbook

    workbook = load_workbook(
        BytesIO(content),
        data_only=False,
        read_only=True,
    )
    try:
        worksheet = workbook[sheet_name]
        if worksheet.max_row - 1 > MAX_ROWS:
            raise ValueError(
                f"Uploaded table exceeds the maximum of {MAX_ROWS:,} rows."
            )
        if worksheet.max_column > MAX_COLUMNS:
            raise ValueError(
                f"Uploaded table exceeds the maximum of {MAX_COLUMNS:,} columns."
            )
        header_values = [
            cell.value
            for cell in next(
                worksheet.iter_rows(min_row=1, max_row=1),
                (),
            )
        ]
        header_names = [
            "" if value is None else str(value).strip()
            for value in header_values
        ]
        if any(not value for value in header_names):
            raise ValueError("Every column must have a non-empty header.")
        if len(set(header_names)) != len(header_names):
            raise ValueError("Duplicate column headers are not supported.")
        if any(len(value) > MAX_COLUMN_NAME_LENGTH for value in header_names):
            raise ValueError(
                f"Column names cannot exceed {MAX_COLUMN_NAME_LENGTH} characters."
            )
        result = dataframe.copy()
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type != "f":
                    continue
                if cell.row == 1:
                    raise ValueError(
                        "Formula cells are not supported in Excel column headers."
                    )
                row_index = cell.row - 2
                column_index = cell.column - 1
                if row_index >= len(result) or column_index >= len(result.columns):
                    continue
                result.iat[row_index, column_index] = str(cell.value)
        return result
    finally:
        workbook.close()


def _is_identifier_column(column_name: str) -> bool:
    tokens = {token for token in re.split(r"[^a-z0-9]+", column_name.lower()) if token}
    return bool(tokens & ID_NAME_TOKENS)


def _has_lossy_numeric_shape(value: str) -> bool:
    unsigned = value.lstrip("+-")
    integer_part = unsigned.split(".", 1)[0]
    if len(integer_part) > 1 and integer_part.startswith("0"):
        return True
    if re.fullmatch(r"[+-]?\d+", value):
        try:
            return abs(int(value)) > SAFE_INTEGER_LIMIT
        except ValueError:
            return True
    return False


def _format_name(extension: str) -> str:
    return {
        ".csv": "CSV",
        ".tsv": "TSV",
        ".txt": "TXT",
        ".xlsx": "Excel",
    }[extension]
